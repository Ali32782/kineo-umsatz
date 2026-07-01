"""Import Abwesenheiten aus HR-Excel (Mitarbeitende, Art, Von, Bis, Halber Tag)."""
from __future__ import annotations

import calendar
import io
import re
import unicodedata
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

FERIEN_ARTEN = {
    "urlaub",
    "umzug",
    "gleitzeitsaldo bezug",
    "gleitzeit",
    "ferien",
    "bezug gleitzeitsaldo",
}
KRANK_ARTEN = {"krankheit", "krank", "unfall", "unfall krankheit"}

# Vollständige Excel-Namen → MA-Kürzel (falls Auto-Match scheitert)
EXCEL_NAME_ALIASES = {
    "eva monika danko": "Eva.D",
    "valerio lo sasso": "Valerio.S",
    "sonia montero cuevas": "Sonia.M",
    "joelle ramseier": "Joëlle.R",
    "joëlle ramseier": "Joëlle.R",
    "andrina kumin": "Andrina.K",
    "andrina kümin": "Andrina.K",
}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()[:10]
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _is_half_day(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower()
    return s not in ("", "-", "nein", "no", "false", "0")


def weekdays_in_month(start: date, end: date, year: int, month: int) -> float:
    """Arbeitstage (Mo–Fr) im Schnitt [von, bis] ∩ Zielmonat."""
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    start = max(start, month_start)
    end = min(end, month_end)
    if start > end:
        return 0.0
    total = 0.0
    d = start
    while d <= end:
        if d.weekday() < 5:
            total += 1.0
        d += timedelta(days=1)
    return total


def build_ma_lookup(mas) -> dict[str, str]:
    """Normalisierter Name → ma.name."""
    lookup: dict[str, str] = {}
    for ma in mas:
        lookup[_norm(ma.name.replace(".", " "))] = ma.name
        if ma.display_name:
            lookup[_norm(ma.display_name.replace(".", ""))] = ma.name
            parts = _norm(ma.display_name).replace(".", "").split()
            if len(parts) >= 2:
                lookup[f"{parts[0]} {parts[1][0]}"] = ma.name
        name_parts = ma.name.split(".")
        if len(name_parts) == 2:
            lookup[_norm(f"{name_parts[0]} {name_parts[1]}")] = ma.name
    lookup.update(EXCEL_NAME_ALIASES)
    return lookup


def match_ma_name(excel_name: str, mas, lookup: dict[str, str] | None = None) -> str | None:
    lookup = lookup or build_ma_lookup(mas)
    key = _norm(excel_name)
    if key in lookup:
        return lookup[key]
    if key in EXCEL_NAME_ALIASES:
        return EXCEL_NAME_ALIASES[key]

    parts = key.split()
    if len(parts) < 2:
        return None
    first, last = parts[0], parts[-1]

    for ma in mas:
        disp = _norm((ma.display_name or "").replace(".", ""))
        disp_parts = disp.split()
        if not disp_parts:
            continue
        ma_first = disp_parts[0]
        ma_last = disp_parts[-1] if len(disp_parts) > 1 else ""
        ma_last_init = ma_last[0] if ma_last else ""
        name_last = ma.name.split(".")[-1].lower() if "." in ma.name else ""

        if ma_first == first and (
            last.startswith(ma_last_init)
            or (name_last and last.startswith(name_last[0]))
            or (len(ma_last) > 1 and last.startswith(ma_last))
        ):
            return ma.name
    return None


def classify_art(art: str) -> str | None:
    a = _norm(art)
    if a in KRANK_ARTEN or "krank" in a:
        return "krank_t"
    if a in FERIEN_ARTEN or "urlaub" in a or "gleitzeit" in a or "umzug" in a:
        return "ferien_t"
    return None


def parse_abwesenheiten_xlsx(
    content: bytes,
    year: int,
    month: int,
    mas,
) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    lookup = build_ma_lookup(mas)

    by_ma: dict[str, dict[str, float]] = {}
    details: list[dict] = []
    unmatched: list[str] = []
    skipped: list[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        excel_name = ws.cell(row_idx, 1).value
        art_raw = ws.cell(row_idx, 2).value
        von = _parse_date(ws.cell(row_idx, 3).value)
        bis = _parse_date(ws.cell(row_idx, 4).value)
        halber = _is_half_day(ws.cell(row_idx, 5).value)

        if not excel_name or not art_raw or not von or not bis:
            continue

        field = classify_art(str(art_raw))
        if not field:
            skipped.append({"row": row_idx, "name": str(excel_name), "art": str(art_raw)})
            continue

        ma_name = match_ma_name(str(excel_name), mas, lookup)
        if not ma_name:
            unmatched.append(str(excel_name).strip())
            continue

        days = weekdays_in_month(von, bis, year, month)
        if halber:
            days *= 0.5
        if days <= 0:
            continue

        by_ma.setdefault(ma_name, {"ferien_t": 0.0, "krank_t": 0.0})
        by_ma[ma_name][field] += days

        details.append({
            "ma_name": ma_name,
            "excel_name": str(excel_name).strip(),
            "art": str(art_raw).strip(),
            "von": von.isoformat(),
            "bis": bis.isoformat(),
            "days": days,
            "field": field,
        })

    # Runden auf 0.5
    for ma_name, vals in by_ma.items():
        vals["ferien_t"] = round(vals["ferien_t"] * 2) / 2
        vals["krank_t"] = round(vals["krank_t"] * 2) / 2

    return {
        "by_ma": by_ma,
        "details": details,
        "unmatched": sorted(set(unmatched)),
        "skipped": skipped,
    }
