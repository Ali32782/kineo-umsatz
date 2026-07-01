from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
import os

from database import UmsatzData, MonthlyInput, MAStammdaten
from calc import compute_zeg, compute_soll_tage, zeg_color, MONTH_NAMES_DE, is_employed_in_month
from schedule_utils import get_schedule_entries_for_month
from standort_calc import expand_ma_standort_rows, aggregate_team_summary
from umsatz_agg import monthly_and_year_totals

TEAL="006B6B"; WHITE="FFFFFF"; GRAY="F5F5F5"; DARK="1A1A1A"
EXPORTS_DIR=os.environ.get("EXPORTS_DIR",os.path.join(os.path.dirname(__file__),"../exports"))

def bdr():
    s=Side(style='thin',color='CCCCCC'); return Border(left=s,right=s,top=s,bottom=s)
def fc(c,bold=False,bg=WHITE,fg=DARK,align='center',fmt=None,sz=9,wrap=False):
    c.font=Font(name='Arial',bold=bold,color=fg,size=sz)
    c.fill=PatternFill('solid',start_color=bg)
    c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    c.border=bdr()
    if fmt: c.number_format=fmt
def zeg_bg(v):
    if v is None: return GRAY
    if v>=1.0: return "E8F8E8"
    if v>=0.85: return "FFF8E0"
    return "FFE8E8"

def _active_mas_for_year(db, year):
    all_mas = db.query(MAStammdaten).all()
    return [
        m for m in all_mas
        if any(is_employed_in_month(m.eintritt, m.austritt, year, mo, m.is_active) for mo in range(1, 13))
    ]

def generate_excel(year:int, db:Session)->str:
    mas = _active_mas_for_year(db, year)
    umsatz_all={(r.ma_name,r.month):r.umsatz for r in db.query(UmsatzData).filter(UmsatzData.year==year).all()}
    inputs_all={(r.ma_name,r.month):r for r in db.query(MonthlyInput).filter(MonthlyInput.year==year).all()}
    wb=Workbook(); wb.remove(wb.active)

    # ── Übersicht ────────────────────────────────────────────────────────
    ws=wb.create_sheet("Übersicht")
    total_cols=2+12*2
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c=ws['A1']; c.value=f'KINEO AG  |  Umsatzanalyse {year}'
    c.font=Font(name='Arial',bold=True,color=WHITE,size=13)
    c.fill=PatternFill('solid',start_color=TEAL)
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=28

    ws.row_dimensions[2].height=18
    for col,txt in [(1,'Mitarbeiter/in'),(2,'BG%')]:
        c=ws.cell(row=2,column=col,value=txt)
        fc(c,bold=True,bg=TEAL,fg=WHITE,align='left' if col==1 else 'center')
    for mi,mn in enumerate(range(1,13)):
        sc=3+mi*2
        ws.merge_cells(start_row=2,start_column=sc,end_row=2,end_column=sc+1)
        c=ws.cell(row=2,column=sc,value=MONTH_NAMES_DE[mn][:3])
        c.font=Font(name='Arial',bold=True,color=WHITE,size=9)
        c.fill=PatternFill('solid',start_color=TEAL)
        c.alignment=Alignment(horizontal='center',vertical='center')

    ws.row_dimensions[3].height=16
    fc(ws.cell(row=3,column=1),bold=True,bg=TEAL,fg=WHITE)
    fc(ws.cell(row=3,column=2),bold=True,bg=TEAL,fg=WHITE)
    for mi in range(12):
        for d,txt in enumerate(['Umsatz','ZEG-B']):
            fc(ws.cell(row=3,column=3+mi*2+d,value=txt),bold=True,bg=TEAL,fg=WHITE,sz=8)

    row=4
    for i,ma in enumerate(mas):
        ws.row_dimensions[row].height=17
        bg=WHITE if i%2==0 else GRAY
        fc(ws.cell(row=row,column=1,value=ma.display_name),bold=True,bg=bg,align='left')
        fc(ws.cell(row=row,column=2,value=ma.bg_pct),bg=bg,fmt='0%')
        for mi,mn in enumerate(range(1,13)):
            sc=3+mi*2
            if not is_employed_in_month(ma.eintritt, ma.austritt, year, mn, ma.is_active):
                for d in range(2): fc(ws.cell(row=row,column=sc+d),bg=bg,fg='CCCCCC')
                continue
            umsatz=umsatz_all.get((ma.name,mn),0)
            inp=inputs_all.get((ma.name,mn))
            soll=compute_soll_tage(ma.name,year,mn,db=db)
            if soll==0 and umsatz==0:
                for d in range(2): fc(ws.cell(row=row,column=sc+d),bg=bg,fg='CCCCCC')
                continue
            zeg=compute_zeg(ma.name,year,mn,umsatz,
                ferien_t=inp.ferien_t if inp else 0,kurs_h=inp.kurs_h if inp else 0,
                workshop_h=inp.workshop_h if inp else 0,marketing_h=inp.marketing_h if inp else 0,
                laufanalyse_h=inp.laufanalyse_h if inp else 0,krank_t=inp.krank_t if inp else 0,db=db)
            fc(ws.cell(row=row,column=sc,value=umsatz),bg=bg,fmt="#'##0",align='right')
            zb=zeg['zeg_b']
            fc(ws.cell(row=row,column=sc+1,value=zb),bold=bool(zb),bg=zeg_bg(zb),fmt='0.0%' if zb else None)
        row+=1

    monthly_totals, year_total = monthly_and_year_totals(umsatz_all, mas, year)
    ws.row_dimensions[row].height=18
    fc(ws.cell(row=row,column=1,value='Monatssumme'),bold=True,bg=GRAY,align='left')
    fc(ws.cell(row=row,column=2,value=''),bg=GRAY)
    for mi, total in enumerate(monthly_totals):
        sc=3+mi*2
        fc(ws.cell(row=row,column=sc,value=total),bold=True,bg=GRAY,fmt="#'##0",align='right')
        fc(ws.cell(row=row,column=sc+1,value=''),bg=GRAY)

    ws.column_dimensions['A'].width=16; ws.column_dimensions['B'].width=6
    for mi in range(12):
        ws.column_dimensions[get_column_letter(3+mi*2)].width=10
        ws.column_dimensions[get_column_letter(4+mi*2)].width=8
    ws.freeze_panes='C4'

    # ── Standorte pro Monat ───────────────────────────────────────────────
    ws_s = wb.create_sheet("Standorte")
    ws_s.merge_cells('A1:F1')
    c=ws_s['A1']; c.value=f'Standort-Umsatz & ZEG-B {year}'
    c.font=Font(name='Arial',bold=True,color=WHITE,size=12)
    c.fill=PatternFill('solid',start_color=TEAL)
    c.alignment=Alignment(horizontal='center',vertical='center')
    hdrs=['Monat','Standort','Umsatz CHF','FTE','ZEG-B Ø','Office?']
    for col,h in enumerate(hdrs,1):
        fc(ws_s.cell(row=2,column=col,value=h),bold=True,bg=TEAL,fg=WHITE)
    sr=3
    for mn in range(1,13):
        month_mas = [m for m in mas if is_employed_in_month(m.eintritt, m.austritt, year, mn, m.is_active)]
        expanded=[]
        for ma in month_mas:
            umsatz=umsatz_all.get((ma.name,mn),0)
            inp=inputs_all.get((ma.name,mn))
            zeg=compute_zeg(ma.name,year,mn,umsatz,
                ferien_t=inp.ferien_t if inp else 0,kurs_h=inp.kurs_h if inp else 0,
                workshop_h=inp.workshop_h if inp else 0,marketing_h=inp.marketing_h if inp else 0,
                laufanalyse_h=inp.laufanalyse_h if inp else 0,krank_t=inp.krank_t if inp else 0,db=db)
            row_data={"name":ma.name,"display_name":ma.display_name,"team":ma.team,"umsatz":umsatz,**zeg}
            sched=get_schedule_entries_for_month(ma.name,year,mn,db)
            expanded.extend(expand_ma_standort_rows(row_data,ma.bg_pct,ma.team,sched))
        summary=aggregate_team_summary(expanded)
        for standort, stats in sorted(summary.items()):
            bg=WHITE if sr%2==0 else GRAY
            fc(ws_s.cell(row=sr,column=1,value=MONTH_NAMES_DE[mn][:3]),bg=bg)
            fc(ws_s.cell(row=sr,column=2,value=standort),bg=bg,align='left')
            fc(ws_s.cell(row=sr,column=3,value=stats['umsatz']),bg=bg,fmt="#'##0",align='right')
            fc(ws_s.cell(row=sr,column=4,value=stats.get('fte',0)),bg=bg,fmt='0.0')
            zb=stats.get('zeg_b_avg')
            fc(ws_s.cell(row=sr,column=5,value=zb),bg=zeg_bg(zb),fmt='0.0%' if zb else None)
            fc(ws_s.cell(row=sr,column=6,value='Ja' if stats.get('is_office') else ''),bg=bg)
            sr+=1

    # ── Monthly sheets ───────────────────────────────────────────────────
    for mn in range(1,13):
        ws_m=wb.create_sheet(f"{MONTH_NAMES_DE[mn][:3]} {year}")
        ws_m.merge_cells('A1:R1')
        c=ws_m['A1']; c.value=f'KINEO AG  |  {MONTH_NAMES_DE[mn]} {year}'
        c.font=Font(name='Arial',bold=True,color=WHITE,size=11)
        c.fill=PatternFill('solid',start_color=TEAL)
        c.alignment=Alignment(horizontal='center',vertical='center')
        ws_m.row_dimensions[1].height=24

        hdrs=['Name','BG%','Soll','Ferien','Kurse(h)','Workshop(h)','Mkt(h)','Lauf(h)',
              'Mgmt(T)','Leit(T)','Krank(T)','Prod-A','Prod-B','Prod-C','Umsatz','ZEG-A','ZEG-B','ZEG-C']
        ws_m.row_dimensions[2].height=28
        for col,h in enumerate(hdrs,1):
            fc(ws_m.cell(row=2,column=col,value=h),bold=True,bg=TEAL,fg=WHITE,sz=8,wrap=True)

        month_mas = [m for m in mas if is_employed_in_month(m.eintritt, m.austritt, year, mn, m.is_active)]
        for i,ma in enumerate(month_mas):
            r=i+3; ws_m.row_dimensions[r].height=17
            bg=WHITE if i%2==0 else GRAY
            umsatz=umsatz_all.get((ma.name,mn),0)
            inp=inputs_all.get((ma.name,mn))
            soll=compute_soll_tage(ma.name,year,mn,db=db)
            zeg=compute_zeg(ma.name,year,mn,umsatz,
                ferien_t=inp.ferien_t if inp else 0,kurs_h=inp.kurs_h if inp else 0,
                workshop_h=inp.workshop_h if inp else 0,marketing_h=inp.marketing_h if inp else 0,
                laufanalyse_h=inp.laufanalyse_h if inp else 0,krank_t=inp.krank_t if inp else 0,db=db)
            vals=[ma.display_name,ma.bg_pct,soll,
                inp.ferien_t if inp else None,inp.kurs_h if inp else None,
                inp.workshop_h if inp else None,inp.marketing_h if inp else None,
                inp.laufanalyse_h if inp else None,
                zeg['mgmt_t'] or None,zeg['leit_t'] or None,inp.krank_t if inp else None,
                zeg['prod_a'],zeg['prod_b'],zeg['prod_c'],
                umsatz,zeg['zeg_a'],zeg['zeg_b'],zeg['zeg_c']]
            fmts=['','0%','0.0','0.0','0.0','0.0','0.0','0.0','0.00','0.00','0.0',
                  '0.0','0.0','0.0',"#'##0",'0.0%','0.0%','0.0%']
            for col,(val,fmt) in enumerate(zip(vals,fmts),1):
                cbg=bg
                if col in (16,17,18) and isinstance(val,float): cbg=zeg_bg(val)
                fc(ws_m.cell(row=r,column=col,value=val),bold=(col==1),bg=cbg,
                   align='left' if col==1 else 'center',fmt=fmt if val else None)

        for ci,w in enumerate([16,6,7,7,7,8,7,7,7,7,7,7,7,7,11,8,8,8],1):
            ws_m.column_dimensions[get_column_letter(ci)].width=w
        ws_m.freeze_panes='A3'

    os.makedirs(EXPORTS_DIR,exist_ok=True)
    path=os.path.join(EXPORTS_DIR,f"Kineo_Umsatzanalyse_{year}.xlsx")
    wb.save(path); return path
