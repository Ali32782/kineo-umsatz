"""Fitness-Abo Excel (Thalwil) → monatsweise Mitgliederzahlen für Ilaria / CC."""
from __future__ import annotations

import io
import re
from datetime import date
from typing import BinaryIO

from openpyxl import load_workbook


_KW_RE = re.compile(r"KW\s*(\d+)", re.IGNORECASE)


def _parse_kw(label) -> int | None:
    if label is None:
        return None
    m = _KW_RE.search(str(label).strip())
    if not m:
        return None
    try:
        kw = int(m.group(1))
    except ValueError:
        return None
    if 1 <= kw <= 53:
        return kw
    return None


def _num(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("*", "").replace("'", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _kw_to_month(year: int, kw: int) -> int | None:
    """ISO-Donnerstag liegt immer im Kalenderjahr der KW-Nummer."""
    try:
        return date.fromisocalendar(year, kw, 4).month
    except ValueError:
        return None


def parse_fitness_abo_excel(source: str | BinaryIO | bytes) -> list[dict]:
    """
    Liest „Mitglieder Gesamt“ je KW und aggregiert auf Monat:
    letzter vorhandener KW-Wert > 0 im Monat.

    Returns: [{year, month, count}, ...]
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row) if row else [])
    wb.close()

    # Letzter KW-Wert pro (year, month)
    last_by_ym: dict[tuple[int, int], float] = {}

    i = 0
    while i < len(rows):
        cell0 = rows[i][0] if rows[i] else None
        year = None
        if isinstance(cell0, int) and 2000 <= cell0 <= 2100:
            year = cell0
        elif isinstance(cell0, float) and cell0 == int(cell0) and 2000 <= int(cell0) <= 2100:
            year = int(cell0)
        if year is None:
            i += 1
            continue

        # Header-Zeile mit KW-Spalten suchen
        header_idx = None
        for j in range(i + 1, min(i + 6, len(rows))):
            labels = rows[j]
            if any(_parse_kw(c) for c in labels[1:]):
                header_idx = j
                break
        if header_idx is None:
            i += 1
            continue

        kws = [_parse_kw(c) for c in rows[header_idx]]
        # Mitglieder-Gesamt-Zeile
        mg_idx = None
        for j in range(header_idx + 1, min(header_idx + 12, len(rows))):
            label = rows[j][0] if rows[j] else None
            if label and "mitglieder gesamt" in str(label).strip().lower():
                mg_idx = j
                break
        if mg_idx is None:
            i = header_idx + 1
            continue

        vals = rows[mg_idx]
        for col, kw in enumerate(kws):
            if kw is None or col >= len(vals):
                continue
            count = _num(vals[col])
            if count is None or count <= 0:
                continue
            month = _kw_to_month(year, kw)
            if month is None:
                continue
            last_by_ym[(year, month)] = count

        i = mg_idx + 1

    out = [
        {"year": y, "month": m, "count": c}
        for (y, m), c in sorted(last_by_ym.items())
    ]
    return out
