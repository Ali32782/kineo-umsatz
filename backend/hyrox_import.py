"""Training-Club Rechnungen → monatsweise HYROX-Umsätze (Nina / Selbstzahler)."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime
from typing import BinaryIO

from openpyxl import load_workbook


def _num(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("'", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _header_map(row: list) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key and key not in out:
            out[key] = i
    return out


def parse_hyrox_invoices_excel(source: str | BinaryIO | bytes) -> list[dict]:
    """
    Liest Training-Club Rechnungen (Sheet «Meine Finanzen»).

    Filter: Rechnungsstatus «Bezahlt», Stunde enthält «Hyrox».
    Aggregation: Summe Total nach Kaufdatum (Jahr/Monat).

    Returns: [{year, month, umsatz, count}, ...]
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) if r else [] for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return []

    hdr = _header_map(rows[0])
    # Fallbacks auf Spaltenindex (0-basiert) laut Export
    i_total = hdr.get("total", 4)
    i_date = hdr.get("kaufdatum", 7)
    i_status = hdr.get("rechnungsstatus", 9)
    i_stunde = hdr.get("stunde", 13)

    by_month: dict[tuple[int, int], float] = defaultdict(float)
    counts: dict[tuple[int, int], int] = defaultdict(int)

    for row in rows[1:]:
        if not row:
            continue
        status_raw = row[i_status] if i_status < len(row) else None
        status = str(status_raw or "").strip().lower()
        if status != "bezahlt":
            continue
        stunde_raw = row[i_stunde] if i_stunde < len(row) else None
        stunde = str(stunde_raw or "").strip().lower()
        if "hyrox" not in stunde:
            continue
        d = _parse_date(row[i_date] if i_date < len(row) else None)
        amt = _num(row[i_total] if i_total < len(row) else None)
        if not d or amt is None:
            continue
        key = (d.year, d.month)
        by_month[key] += amt
        counts[key] += 1

    out = []
    for (year, month) in sorted(by_month):
        umsatz = round(by_month[(year, month)], 2)
        if umsatz <= 0:
            continue
        out.append({
            "year": year,
            "month": month,
            "umsatz": umsatz,
            "count": counts[(year, month)],
        })
    return out
