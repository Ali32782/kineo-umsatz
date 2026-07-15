"""Runnerslab Umsatz-Excel → monatsweise Umsätze für Marc / CC."""
from __future__ import annotations

import io
from typing import BinaryIO

from openpyxl import load_workbook


def _num(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("'", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _year_cell(val) -> int | None:
    if isinstance(val, int) and 2000 <= val <= 2100:
        return val
    if isinstance(val, float) and val == int(val) and 2000 <= int(val) <= 2100:
        return int(val)
    return None


def parse_runnerslab_excel(source: str | BinaryIO | bytes) -> list[dict]:
    """
    Liest Zeile „Total“ je Jahresblock (Spalten Jan–Dez).

    Returns: [{year, month, umsatz}, ...] nur Monate mit umsatz > 0.
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row) if row else [])
    wb.close()

    out: list[dict] = []
    i = 0
    while i < len(rows):
        year = _year_cell(rows[i][0] if rows[i] else None)
        if year is None:
            i += 1
            continue

        total_idx = None
        for j in range(i + 1, min(i + 8, len(rows))):
            label = rows[j][0] if rows[j] else None
            if label and str(label).strip().lower() == "total":
                total_idx = j
                break
        if total_idx is None:
            i += 1
            continue

        vals = rows[total_idx]
        # Spalten 2–13 = Jan–Dez (1-basiert col 2 → index 1)
        for month in range(1, 13):
            col = month  # index: Jan=1
            if col >= len(vals):
                break
            umsatz = _num(vals[col])
            if umsatz is None or umsatz <= 0:
                continue
            out.append({"year": year, "month": month, "umsatz": round(umsatz, 2)})

        i = total_idx + 1

    return out
