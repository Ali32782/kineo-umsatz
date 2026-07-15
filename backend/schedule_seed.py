"""Arbeitspläne aus Kineo_Standort_Uebersicht_2026.xlsx laden."""
from __future__ import annotations

import math
import os
from pathlib import Path

from calc import HALBTAG_PCT, MA_PATTERNS, day_pct_to_halves, schedule_needs_reseed

DATA_DIR = Path(__file__).resolve().parent / "data"
DEFAULT_XLSX = DATA_DIR / "Kineo_Standort_Uebersicht_2026.xlsx"

EXCEL_NAME_TO_MA = {
    "Andrina K.": "Andrina.K",
    "Barbara V.": "Barbara.V",
    "Carmen W.": "Carmen.W",
    "Clara B.": "Clara.B",
    "Emma L.": "Emma.L",
    "Eva D.": "Eva.D",
    "Hanna R.": "Hanna.R",
    "Helen S.": "Helen.S",
    "Joëlle R.": "Joëlle.R",
    "Lucrecia G.": "Lucrecia.G",
    "Martino C.": "Martino.C",
    "Meike V.": "Meike.V",
    "Noah S.": "Noah.S",
    "Pablo G.": "Pablo.G",
    "Pablo M.": "Pablo.M",
    "Raphael H.": "Raphael.H",
    "Sereina U.": "Sereina.U",
    "Sonia M.": "Sonia.M",
    "Valerio S.": "Valerio.S",
}

DAY_COLS = ["Mo", "Di", "Mi", "Do", "Fr"]
WEEKDAY_KEYS = ["mo", "di", "mi", "do", "fr"]

STANDORT_ALIASES = {
    "stauf.": "Stauffacher",
    "stauffacher": "Stauffacher",
    "off.": "Management",
    "office": "Management",
    "management": "Management",
    "wipkingen": "Wipkingen",
    "thalwil": "Thalwil",
    "escher wyss": "Escher Wyss",
    "seefeld": "Seefeld",
    "zollikon": "Zollikon",
}


def normalize_standort(raw: str) -> str:
    key = raw.strip().lower()
    if key in STANDORT_ALIASES:
        return STANDORT_ALIASES[key]
    return raw.strip()


def _is_blank(cell) -> bool:
    if cell is None:
        return True
    if isinstance(cell, float) and math.isnan(cell):
        return True
    return not str(cell).strip()


def parse_day_cell(cell, day_pct: float) -> tuple[float, str | None, float, str | None]:
    """Excel-Zelle + MA_PATTERNS-Tagesanteil → (vm_pct, vm_standort, nm_pct, nm_standort)."""
    if _is_blank(cell):
        return 0.0, None, 0.0, None

    cell = str(cell).strip()

    if cell.upper().endswith(" VM"):
        loc = normalize_standort(cell[:-3].strip())
        return HALBTAG_PCT, loc, 0.0, None

    if "/" in cell:
        vm_raw, nm_raw = cell.split("/", 1)
        return (
            HALBTAG_PCT,
            normalize_standort(vm_raw),
            HALBTAG_PCT,
            normalize_standort(nm_raw),
        )

    loc = normalize_standort(cell)
    vm, nm = day_pct_to_halves(day_pct)
    if day_pct <= HALBTAG_PCT + 0.001:
        return vm, loc, 0.0, None
    return vm, loc, nm, loc


def load_excel_schedules(xlsx_path: Path | None = None) -> dict[str, list[dict]]:
    """Alle Arbeitspläne aus der Standort-Übersicht lesen."""
    from openpyxl import load_workbook

    path = xlsx_path or Path(os.environ.get("SCHEDULE_XLSX", DEFAULT_XLSX))
    if not path.exists():
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    schedules: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        display = str(row[0]).strip()
        if display not in EXCEL_NAME_TO_MA:
            continue
        ma_name = EXCEL_NAME_TO_MA[display]
        pat = MA_PATTERNS.get(ma_name, {})
        days: list[dict] = []

        for wd in range(5):
            cell = row[2 + wd] if len(row) > 2 + wd else None
            day_pct = pat.get(WEEKDAY_KEYS[wd], 0) or 0
            vm, vm_loc, nm, nm_loc = parse_day_cell(cell, day_pct)
            if vm or nm:
                days.append({
                    "weekday": wd,
                    "vm_pct": vm,
                    "vm_standort": vm_loc,
                    "nm_pct": nm,
                    "nm_standort": nm_loc,
                })

        schedules[ma_name] = days

    wb.close()
    return schedules


def _day_matches(entry, day: dict) -> bool:
    if not entry:
        return False
    return (
        round(entry.vm_pct or 0, 2) == round(day.get("vm_pct") or 0, 2)
        and round(entry.nm_pct or 0, 2) == round(day.get("nm_pct") or 0, 2)
        and (entry.vm_standort or None) == day.get("vm_standort")
        and (entry.nm_standort or None) == day.get("nm_standort")
    )


def schedule_needs_excel_reseed(entries, excel_days: list[dict]) -> bool:
    if schedule_needs_reseed(entries):
        return True
    if not entries and excel_days:
        return True
    by_wd = {e.weekday: e for e in entries}
    excel_by_wd = {d["weekday"]: d for d in excel_days}
    if set(by_wd) != set(excel_by_wd):
        return True
    for wd, day in excel_by_wd.items():
        if not _day_matches(by_wd.get(wd), day):
            return True
    return False
